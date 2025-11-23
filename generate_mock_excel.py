"""
Script to generate a mock Excel file with example location data for PV, Storage, and Charging products.
"""

import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Sample addresses in Germany
ADDRESSES = [
    "Münchner Straße 15, 80331 München",
    "Hauptstraße 42, 10115 Berlin",
    "Königsallee 1, 40212 Düsseldorf",
    "Neuer Wall 10, 20354 Hamburg",
    "Zeil 5, 60313 Frankfurt am Main",
    "Königstraße 30, 70173 Stuttgart",
    "Breite Straße 8, 50667 Köln",
    "Marktplatz 2, 04109 Leipzig",
    "Bahnhofstraße 20, 01099 Dresden",
    "Lange Straße 1, 30159 Hannover",
    "Kaiserstraße 50, 76133 Karlsruhe",
    "Friedrichstraße 123, 10117 Berlin",
    "Schadowstraße 11, 40212 Düsseldorf",
    "Ballindamm 1, 20095 Hamburg",
    "Goethestraße 7, 60313 Frankfurt am Main",
    "Calwer Straße 45, 70173 Stuttgart",
    "Hohe Straße 100, 50667 Köln",
    "Grimmaische Straße 1, 04109 Leipzig",
    "Prager Straße 1, 01069 Dresden",
    "Georgstraße 50, 30159 Hannover",
]

# Sample location names
PV_NAMES = [
    "Gewerbegebiet Nord - Lagerhalle A",
    "Industriezentrum Süd - Produktionshalle",
    "Einkaufszentrum City - Dachfläche",
    "Bürokomplex Mitte - Hauptgebäude",
    "Logistikzentrum West - Halle 3",
    "Produktionsstätte Ost - Werk 1",
    "Verwaltungsgebäude Zentrum - Hauptsitz",
    "Einzelhandelszentrum - Filiale Nord",
    "Gewerbehof - Gebäude B",
    "Industriepark - Halle 5",
    "Büropark - Gebäude 2",
    "Einkaufszentrum - Erweiterung",
    "Lagerhalle - Standort 3",
    "Produktionshalle - Linie 2",
    "Verwaltungsgebäude - Nebengebäude",
]

STORAGE_NAMES = [
    "Bürokomplex Hauptstraße - Hauptgebäude",
    "Produktionsstätte Industriegebiet - Werk 2",
    "Einkaufszentrum City - Hauptgebäude",
    "Verwaltungszentrum - Gebäude A",
    "Gewerbegebiet - Halle 1",
    "Industriezentrum - Produktionshalle B",
    "Büropark - Hauptgebäude",
    "Logistikzentrum - Verwaltungsgebäude",
    "Einzelhandelszentrum - Hauptgebäude",
    "Produktionsstätte - Verwaltungsgebäude",
    "Gewerbehof - Hauptgebäude",
    "Industriepark - Verwaltungsgebäude",
    "Bürokomplex - Nebengebäude",
    "Einkaufszentrum - Verwaltungsgebäude",
    "Lagerhalle - Bürogebäude",
]

CHARGING_NAMES = [
    "Parkhaus Innenstadt - Ebene 1",
    "Einkaufszentrum - Parkplatz Nord",
    "Bürokomplex - Tiefgarage",
    "Bahnhof - Parkplatz P1",
    "Flughafen - Parkhaus Terminal 1",
    "Krankenhaus - Parkplatz Haupteingang",
    "Hotel City Center - Tiefgarage",
    "Restaurant Mile - Parkplatz",
    "Fitnessstudio - Parkplatz",
    "Supermarkt - Kundenparkplatz",
    "Kino - Parkplatz",
    "Museum - Parkplatz",
    "Stadion - Parkplatz A",
    "Universität - Parkplatz Campus",
    "Verwaltungsgebäude - Besucherparkplatz",
]


INDUSTRIES = [
    "Energie & Versorgung",
    "Produktion & Fertigung",
    "Logistik & Transport",
    "Einzelhandel & Handel",
    "Gastronomie & Hotellerie",
    "Büro & Verwaltung",
    "Gesundheitswesen",
    "Bildung & Forschung",
    "Immobilien & Bau",
    "IT & Telekommunikation",
    "Automobil & Mobilität",
    "Chemie & Pharma",
    "Lebensmittel & Getränke",
    "Textil & Mode",
    "Maschinenbau",
    "Elektronik & Elektrotechnik",
    "Landwirtschaft",
    "Sonstige"
]

def generate_pv_data(num_rows=20):
    """Generate PV location data"""
    data = []
    for i in range(1, num_rows + 1):
        data.append({
            "location_id": i,
            "location_name": random.choice(PV_NAMES),
            "address": random.choice(ADDRESSES),
            "product": "pv",
            "eigentuemer": random.choice(["Ja", "Nein"]),
            "umsatz": random.randint(100000, 100000000),
            "mitarbeiterzahl": random.randint(1, 10000),
            "branche": random.choice(INDUSTRIES),
            "roof_area_sqm": random.randint(100, 2000),
            "solar_irradiation": random.randint(900, 1250),
            "roof_orientation_degrees": random.choice([150, 160, 170, 180, 190, 200, 210]),
            "roof_tilt_degrees": random.randint(20, 45),
            "electricity_price_eur": round(random.uniform(0.25, 0.45), 2),
        })
    return data


def generate_storage_data(num_rows=20):
    """Generate Storage location data"""
    data = []
    for i in range(1, num_rows + 1):
        data.append({
            "location_id": i,
            "location_name": random.choice(STORAGE_NAMES),
            "address": random.choice(ADDRESSES),
            "product": "storage",
            "eigentuemer": random.choice(["Ja", "Nein"]),
            "umsatz": random.randint(100000, 100000000),
            "mitarbeiterzahl": random.randint(1, 10000),
            "branche": random.choice(INDUSTRIES),
            "existing_pv_kwp": random.randint(50, 300),
            "annual_consumption_kwh": random.randint(50000, 400000),
            "peak_load_kw": random.randint(50, 300),
            "grid_connection_kw": random.randint(50, 400),
            "electricity_price_eur": round(random.uniform(0.25, 0.45), 2),
        })
    return data


def generate_charging_data(num_rows=20):
    """Generate Charging location data"""
    data = []
    for i in range(1, num_rows + 1):
        data.append({
            "location_id": i,
            "location_name": random.choice(CHARGING_NAMES),
            "address": random.choice(ADDRESSES),
            "product": "charging",
            "eigentuemer": random.choice(["Ja", "Nein"]),
            "umsatz": random.randint(100000, 100000000),
            "mitarbeiterzahl": random.randint(1, 10000),
            "branche": random.choice(INDUSTRIES),
            "parking_spaces": random.randint(20, 300),
            "daily_traffic_volume": random.randint(200, 5000),
            "avg_parking_duration_min": random.randint(30, 240),
            "grid_connection_kw": random.randint(50, 500),
            "ev_density_percent": round(random.uniform(2, 20), 1),
        })
    return data


def create_excel_file(filename="mock_locations.xlsx"):
    """Create Excel file with three sheets"""
    
    # Create dataframes
    pv_df = pd.DataFrame(generate_pv_data())
    storage_df = pd.DataFrame(generate_storage_data())
    charging_df = pd.DataFrame(generate_charging_data())
    
    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write each sheet
        pv_df.to_excel(writer, sheet_name='PV', index=False)
        storage_df.to_excel(writer, sheet_name='Storage', index=False)
        charging_df.to_excel(writer, sheet_name='Charging', index=False)
    
    # Format the Excel file
    wb = load_workbook(filename)
    
    # Format each sheet
    for sheet_name in ['PV', 'Storage', 'Charging']:
        ws = wb[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Center align numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze first row
        ws.freeze_panes = "A2"
    
    wb.save(filename)
    print(f"Excel file created: {filename}")
    print(f"   - PV sheet: {len(pv_df)} rows")
    print(f"   - Storage sheet: {len(storage_df)} rows")
    print(f"   - Charging sheet: {len(charging_df)} rows")


if __name__ == "__main__":
    create_excel_file("mock_locations.xlsx")

